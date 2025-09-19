from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Form, Depends
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import aiofiles
import tempfile
import re
import PyPDF2
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from emergentintegrations.llm.chat import LlmChat, UserMessage


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

logger = logging.getLogger(__name__)

# Pydantic Models
class Student(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    roll_number: str

class AssignmentSubmission(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    student_name: str
    roll_number: str
    file_name: str
    original_text: str
    marks: Optional[int] = None
    max_marks: Optional[int] = None
    feedback: Optional[str] = None
    submitted_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    evaluated_at: Optional[datetime] = None
    evaluated_by: Optional[str] = None

class AssignmentSubmissionCreate(BaseModel):
    student_name: str
    roll_number: str
    file_name: str
    original_text: str

class EvaluationRequest(BaseModel):
    submission_id: str
    max_marks: int
    evaluator_name: str = "Admin"

class Admin(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password_hash: str
    full_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LoginRequest(BaseModel):
    username: str
    password: str

# Initialize LLM Chat
def get_llm_chat():
    return LlmChat(
        api_key=os.environ.get('EMERGENT_LLM_KEY'),
        session_id=str(uuid.uuid4()),
        system_message="""You are an experienced university professor evaluating student assignments. 
        Your task is to:
        1. Carefully read and analyze the assignment content
        2. Provide marks based on the maximum marks available
        3. Give constructive feedback highlighting strengths and areas for improvement
        4. Be fair but thorough in your evaluation
        5. Consider content quality, understanding, presentation, and completeness"""
    ).with_model("openai", "gpt-4o")

# Utility Functions
def extract_text_from_pdf(file_content: bytes) -> str:
    try:
        pdf_file = io.BytesIO(file_content)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from PDF: {e}")
        return ""

def extract_text_from_docx(file_content: bytes) -> str:
    try:
        docx_file = io.BytesIO(file_content)
        doc = Document(docx_file)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error extracting text from DOCX: {e}")
        return ""

def extract_student_details(text: str) -> tuple[str, str]:
    """Extract student name and roll number from text using patterns"""
    name = "Unknown Student"
    roll_number = "Unknown"
    
    # Common patterns for student names
    name_patterns = [
        r"(?:Name|Student Name|Full Name)[:=\s]*([A-Za-z\s]+?)(?:\n|Roll|ID|$)",
        r"Name\s*[:=]\s*([A-Za-z\s]+)",
        r"Student[:=\s]*([A-Za-z\s]+?)(?:\n|Roll|ID)",
        r"^([A-Za-z\s]{2,30})(?:\n|Roll|ID|Reg)",
    ]
    
    # Common patterns for roll numbers
    roll_patterns = [
        r"(?:Roll No|Roll Number|Registration No|Reg No|ID|Student ID)[:=\s]*([A-Za-z0-9\-/]+)",
        r"Roll[:=\s]*([A-Za-z0-9\-/]+)",
        r"(?:ID|REG)[:=\s]*([A-Za-z0-9\-/]+)",
        r"([0-9]{4,}[A-Za-z0-9\-/]*)",
    ]
    
    # Try to extract name
    for pattern in name_patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if match:
            name = match.group(1).strip()
            if len(name) > 2 and not any(char.isdigit() for char in name):
                break
    
    # Try to extract roll number
    for pattern in roll_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            roll_number = match.group(1).strip()
            break
    
    return name, roll_number

async def create_excel_report():
    """Create Excel report of all submissions"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Assignment Submissions"
    
    # Headers
    headers = ["Student Name", "Roll Number", "File Name", "Marks", "Max Marks", "Percentage", "Feedback", "Submitted Date", "Evaluated Date"]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get all submissions
    submissions = await db.assignments.find().sort("submitted_at", -1).to_list(None)
    
    # Add data
    for row, submission in enumerate(submissions, 2):
        ws.cell(row=row, column=1, value=submission.get("student_name", ""))
        ws.cell(row=row, column=2, value=submission.get("roll_number", ""))
        ws.cell(row=row, column=3, value=submission.get("file_name", ""))
        ws.cell(row=row, column=4, value=submission.get("marks", ""))
        ws.cell(row=row, column=5, value=submission.get("max_marks", ""))
        
        # Calculate percentage
        if submission.get("marks") and submission.get("max_marks"):
            percentage = round((submission["marks"] / submission["max_marks"]) * 100, 2)
            ws.cell(row=row, column=6, value=f"{percentage}%")
        else:
            ws.cell(row=row, column=6, value="")
            
        ws.cell(row=row, column=7, value=submission.get("feedback", ""))
        # Handle datetime formatting - data might be stored as strings
        submitted_at = submission.get("submitted_at", "")
        if submitted_at:
            if isinstance(submitted_at, str):
                try:
                    submitted_at = datetime.fromisoformat(submitted_at).strftime("%Y-%m-%d %H:%M")
                except:
                    submitted_at = submitted_at[:19] if len(submitted_at) > 19 else submitted_at
            else:
                submitted_at = submitted_at.strftime("%Y-%m-%d %H:%M")
        ws.cell(row=row, column=8, value=submitted_at)
        
        evaluated_at = submission.get("evaluated_at", "")
        if evaluated_at:
            if isinstance(evaluated_at, str):
                try:
                    evaluated_at = datetime.fromisoformat(evaluated_at).strftime("%Y-%m-%d %H:%M")
                except:
                    evaluated_at = evaluated_at[:19] if len(evaluated_at) > 19 else evaluated_at
            else:
                evaluated_at = evaluated_at.strftime("%Y-%m-%d %H:%M")
        ws.cell(row=row, column=9, value=evaluated_at)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

# API Routes
@api_router.post("/upload", response_model=AssignmentSubmission)
async def upload_assignment(file: UploadFile = File(...)):
    """Upload and process assignment file"""
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="No file selected")
        
        # Check file type
        allowed_extensions = ['.pdf', '.docx', '.txt']
        file_extension = Path(file.filename).suffix.lower()
        
        if file_extension not in allowed_extensions:
            raise HTTPException(status_code=400, detail="Only PDF, DOCX, and TXT files are allowed")
        
        # Read file content
        file_content = await file.read()
        
        # Extract text based on file type
        if file_extension == '.pdf':
            text = extract_text_from_pdf(file_content)
        elif file_extension == '.docx':
            text = extract_text_from_docx(file_content)
        else:  # .txt
            text = file_content.decode('utf-8', errors='ignore')
        
        if not text.strip():
            raise HTTPException(status_code=400, detail="Could not extract text from the file")
        
        # Extract student details
        student_name, roll_number = extract_student_details(text)
        
        # Create submission
        submission = AssignmentSubmission(
            student_name=student_name,
            roll_number=roll_number,
            file_name=file.filename,
            original_text=text
        )
        
        # Save to database
        submission_dict = submission.dict()
        submission_dict['submitted_at'] = submission_dict['submitted_at'].isoformat()
        await db.assignments.insert_one(submission_dict)
        
        return submission
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing file upload: {e}")
        raise HTTPException(status_code=500, detail="Error processing file")

@api_router.post("/evaluate")
async def evaluate_assignment(request: EvaluationRequest):
    """Evaluate assignment using LLM"""
    try:
        # Get submission from database
        submission = await db.assignments.find_one({"id": request.submission_id})
        if not submission:
            raise HTTPException(status_code=404, detail="Submission not found")
        
        # Prepare evaluation prompt
        evaluation_prompt = f"""
        Please evaluate this student assignment out of {request.max_marks} marks.

        Assignment Content:
        {submission['original_text'][:3000]}  # Limit to avoid token limits

        Please provide:
        1. A numerical score out of {request.max_marks}
        2. Detailed feedback explaining the grade

        Format your response as:
        MARKS: [numerical score]
        FEEDBACK: [detailed feedback]
        """
        
        # Get LLM evaluation
        chat = get_llm_chat()
        user_message = UserMessage(text=evaluation_prompt)
        response = await chat.send_message(user_message)
        
        # Parse response
        marks = None
        feedback = response
        
        # Try to extract marks from response
        marks_match = re.search(r'MARKS?[:=\s]*([0-9]+(?:\.[0-9]+)?)', response, re.IGNORECASE)
        if marks_match:
            marks = int(float(marks_match.group(1)))
            marks = min(marks, request.max_marks)  # Ensure marks don't exceed max
        
        # Extract feedback
        feedback_match = re.search(r'FEEDBACK[:=\s]*(.*)', response, re.IGNORECASE | re.DOTALL)
        if feedback_match:
            feedback = feedback_match.group(1).strip()
        
        # Update submission in database
        update_data = {
            "marks": marks,
            "max_marks": request.max_marks,
            "feedback": feedback,
            "evaluated_at": datetime.now(timezone.utc).isoformat(),
            "evaluated_by": request.evaluator_name
        }
        
        await db.assignments.update_one(
            {"id": request.submission_id},
            {"$set": update_data}
        )
        
        return {
            "success": True,
            "marks": marks,
            "max_marks": request.max_marks,
            "feedback": feedback,
            "message": "Assignment evaluated successfully"
        }
        
    except Exception as e:
        logger.error(f"Error evaluating assignment: {e}")
        raise HTTPException(status_code=500, detail="Error evaluating assignment")

@api_router.get("/submissions", response_model=List[AssignmentSubmission])
async def get_submissions(
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 50
):
    """Get all submissions with optional search"""
    try:
        query = {}
        if search:
            query = {
                "$or": [
                    {"student_name": {"$regex": search, "$options": "i"}},
                    {"roll_number": {"$regex": search, "$options": "i"}},
                    {"file_name": {"$regex": search, "$options": "i"}}
                ]
            }
        
        submissions = await db.assignments.find(query).sort("submitted_at", -1).skip(skip).limit(limit).to_list(None)
        
        # Convert datetime strings back to datetime objects for response
        for submission in submissions:
            if submission.get('submitted_at'):
                submission['submitted_at'] = datetime.fromisoformat(submission['submitted_at'])
            if submission.get('evaluated_at'):
                submission['evaluated_at'] = datetime.fromisoformat(submission['evaluated_at'])
        
        return [AssignmentSubmission(**submission) for submission in submissions]
        
    except Exception as e:
        logger.error(f"Error fetching submissions: {e}")
        raise HTTPException(status_code=500, detail="Error fetching submissions")

@api_router.get("/submission/{submission_id}", response_model=AssignmentSubmission)
async def get_submission(submission_id: str):
    """Get specific submission"""
    try:
        submission = await db.assignments.find_one({"id": submission_id})
        if not submission:
            raise HTTPException(status_code=404, detail="Submission not found")
        
        # Convert datetime strings
        if submission.get('submitted_at'):
            submission['submitted_at'] = datetime.fromisoformat(submission['submitted_at'])
        if submission.get('evaluated_at'):
            submission['evaluated_at'] = datetime.fromisoformat(submission['evaluated_at'])
        
        return AssignmentSubmission(**submission)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching submission: {e}")
        raise HTTPException(status_code=500, detail="Error fetching submission")

@api_router.get("/export/excel")
async def export_excel():
    """Export all submissions to Excel"""
    try:
        excel_buffer = await create_excel_report()
        
        def iter_excel():
            yield excel_buffer.read()
        
        return StreamingResponse(
            io.BytesIO(excel_buffer.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=assignment_submissions.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"Error creating Excel export: {e}")
        raise HTTPException(status_code=500, detail="Error creating Excel export")

@api_router.get("/stats")
async def get_stats():
    """Get dashboard statistics"""
    try:
        total_submissions = await db.assignments.count_documents({})
        evaluated_submissions = await db.assignments.count_documents({"marks": {"$ne": None}})
        pending_submissions = total_submissions - evaluated_submissions
        
        # Get average marks
        pipeline = [
            {"$match": {"marks": {"$ne": None}, "max_marks": {"$ne": None}}},
            {"$group": {
                "_id": None,
                "avg_marks": {"$avg": "$marks"},
                "avg_max_marks": {"$avg": "$max_marks"}
            }}
        ]
        avg_result = await db.assignments.aggregate(pipeline).to_list(1)
        avg_percentage = 0
        if avg_result and avg_result[0]['avg_max_marks'] > 0:
            avg_percentage = (avg_result[0]['avg_marks'] / avg_result[0]['avg_max_marks']) * 100
        
        return {
            "total_submissions": total_submissions,
            "evaluated_submissions": evaluated_submissions,
            "pending_submissions": pending_submissions,
            "average_percentage": round(avg_percentage, 2)
        }
        
    except Exception as e:
        logger.error(f"Error fetching stats: {e}")
        raise HTTPException(status_code=500, detail="Error fetching statistics")

# Simple login endpoint (basic implementation)
@api_router.post("/login")
async def login(request: LoginRequest):
    """Simple login - in production, use proper authentication"""
    # Basic hardcoded admin credentials for MVP
    if request.username == "admin" and request.password == "admin123":
        return {
            "success": True,
            "message": "Login successful",
            "user": {
                "username": "admin",
                "full_name": "Administrator",
                "role": "admin"
            }
        }
    else:
        raise HTTPException(status_code=401, detail="Invalid credentials")

@api_router.get("/")
async def root():
    return {"message": "Smart Assignment Checker API"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()